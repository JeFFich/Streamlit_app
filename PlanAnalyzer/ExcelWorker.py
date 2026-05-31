from openpyxl import load_workbook
from typing import Dict, List, Optional
from collections import defaultdict


class ExcelWorker:
    def __init__(
        self,
        SHEET_PREFIX="TV Flow shorts ",
        LABEL_TOTAL = "Monthly TRPs",
        LABEL_20S="Monthly 20'' TRPs",
        MONTH_ABBR=["Янв", "Фев", "Мар", "Апр", "Май", "Июн",
                    "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]
    ):
        self.SHEET_PREFIX = SHEET_PREFIX
        self.LABEL_TOTAL = LABEL_TOTAL
        self.LABEL_20S = LABEL_20S
        self.MONTH_ABBR = MONTH_ABBR
    
    # =========================================================================
    # ПАРСИНГ EXCEL C ПЛАНОМ
    # =========================================================================
     
    def parse_excel(self, file_obj) -> List[Dict]:
        """
        Парсит Excel-файл с флоучартами.
        Обрабатывает только листы с названием формата "TV Flow charts {Вертикаль}".

        :param file_obj: Файловый объект с Excel-файлом (может быть как путь, так и готовый объект)
        
        :return: Список словарей, каждый описывает одну РК.
        """
        wb = load_workbook(file_obj, data_only=True)
        all_campaigns = []

        for sheet_name in wb.sheetnames:
            # Фильтруем: только листы с нужным префиксом
            if not sheet_name.startswith(self.SHEET_PREFIX):
                continue

            # Извлекаем название вертикали из имени листа
            vertical = sheet_name[len(self.SHEET_PREFIX):].strip()

            ws = wb[sheet_name]
            campaigns_from_sheet = self._parse_vertical_sheet(ws, vertical)
            all_campaigns.extend(campaigns_from_sheet)

        wb.close()
        return all_campaigns

    def _parse_vertical_sheet(self, ws, vertical: str) -> List[Dict]:
        """
        Парсит один лист (одну вертикаль) из Excel.

        Формат листа:
            A1: "Вертикаль: {Название вертикали}"
            Строка 3: шапка (A="Категория"?, B-M=месяцы, N=тип TRP)
            Строки 4+: данные — на каждую категорию 2 строки:
                - строка 1: общий TRP (последний столбец = "Monthly TRPs")
                - строка 2: TRP 20s (последний столбец = "Monthly 20'' TRPs")
            Название категории в столбце A — объединённая ячейка из 2 строк.

        :param ws: Worksheet объект openpyxl.
        :param vertical: Название вертикали.
        :return: Список РК с данного листа.
        """
        campaigns = []

        # Определяем последний столбец с данными (ищем столбец с лейблами TRP)
        # Шапка в строке 3: B-M = месяцы, последний столбец = лейбл
        # Определяем номер последнего столбца по шапке
        label_col = None
        for col in range(2, ws.max_column + 1):
            header_val = ws.cell(row=3, column=col).value
            if header_val is not None and str(header_val).strip() in (self.LABEL_TOTAL, self.LABEL_20S, ""):
                # Ищем столбец N (14-й) или дальше
                pass

        # Более надёжный подход: ищем столбец с лейблами среди данных
        # Проходим по первой строке данных (строка 4) и ищем текстовый лейбл
        label_col = self._find_label_column(ws)

        # Месяцы в столбцах B (2) — M (13)
        month_start_col = 2
        month_end_col = 13  # Включительно

        # Если нашли столбец лейблов, месяцы до него
        if label_col is not None and label_col <= 14:
            month_end_col = label_col - 1

        # Парсим данные начиная со строки 4
        row = 4
        max_row = ws.max_row

        while row <= max_row:
            # Ищем название категории в столбце A
            category_name = ws.cell(row=row, column=1).value

            # Пропускаем пустые строки
            if category_name is None or str(category_name).strip() == "":
                # Возможно, это вторая строка объединённой ячейки — проверяем
                # Попробуем прочитать лейбл из столбца label_col
                label_val = ws.cell(row=row, column=label_col).value if label_col else None
                if label_val is not None and str(label_val).strip() in (self.LABEL_TOTAL, self.LABEL_20S):
                    # Это строка данных без видимого названия (объединённая ячейка)
                    row += 1
                    continue
                row += 1
                continue

            category_name = str(category_name).strip()

            # Определяем, какая из двух строк — total, какая — 20s
            # Читаем лейблы из столбца label_col
            label_row1 = None
            label_row2 = None
            if label_col:
                label_row1 = ws.cell(row=row, column=label_col).value
                label_row2 = ws.cell(row=row + 1, column=label_col).value

            label_row1 = str(label_row1).strip() if label_row1 else ""
            label_row2 = str(label_row2).strip() if label_row2 else ""

            # Определяем, какая строка total, какая 20s
            if self.LABEL_TOTAL in label_row1:
                total_row = row
                s20_row = row + 1
            elif self.LABEL_TOTAL in label_row2:
                total_row = row + 1
                s20_row = row
            else:
                # По умолчанию: первая = total, вторая = 20s
                total_row = row
                s20_row = row + 1

            # Проверяем, что вторая строка существует
            if s20_row > max_row:
                break

            # Читаем значения TRP и цвета по месяцам
            month_data = {}

            for col in range(month_start_col, month_end_col + 1):
                month = col - month_start_col + 1  # 1-12

                trp_total_cell = ws.cell(row=total_row, column=col)
                trp_20s_cell = ws.cell(row=s20_row, column=col)

                trp_total = self._parse_cell_value(trp_total_cell.value)
                trp_20s = self._parse_cell_value(trp_20s_cell.value)

                # Пропускаем пустые месяцы
                if trp_total is None or trp_total == 0:
                    continue

                # Валидация: TRP не может быть отрицательным
                if trp_total < 0:
                    raise ValueError(
                        f"Отрицательный месячный TRP обнаружен: "
                        f"категория='{category_name}', месяц={month}, "
                        f"TRP={trp_total}"
                    )
                if trp_20s is not None and trp_20s < 0:
                    raise ValueError(
                        f"Отрицательный месячный TRP (20s) обнаружен: "
                        f"категория='{category_name}', месяц={month}, "
                        f"TRP_20s={trp_20s}"
                    )

                # Извлекаем цвет ячейки для определения принадлежности к РК
                color = self._get_cell_color(trp_total_cell)

                month_data[month] = {
                    "trp_total": trp_total,
                    "trp_20s": trp_20s if trp_20s is not None else trp_total,
                    "color": color
                }

            # Группируем месяцы по цветам → отдельные РК
            if month_data:
                grouped = self._group_months_by_color(month_data)
                for color, months_info in grouped.items():
                    campaigns.append({
                        "category": category_name,
                        "vertical": vertical,
                        "months": months_info
                    })

            # Переходим к следующей категории (через 2 строки)
            row += 2

        return campaigns

    def _find_label_column(self, ws) -> Optional[int]:
        """
        Находит номер столбца с лейблами типа TRP ("Monthly TRPs" / "Monthly 20'' TRPs").

        Ищет в строках 4-5 (первые строки данных) столбцы после M (13).

        :param ws: Worksheet.
        :return: Номер столбца или None.
        """
        # Ищем в строках 3-6 (шапка и первые данные)
        for check_row in range(3, min(7, ws.max_row + 1)):
            for col in range(13, ws.max_column + 1):
                val = ws.cell(row=check_row, column=col).value
                if val is not None:
                    val_str = str(val).strip()
                    if self.LABEL_TOTAL in val_str or self.LABEL_20S in val_str:
                        return col
        # Если не нашли, предполагаем столбец N (14)
        return 14

    def _parse_cell_value(self, value) -> Optional[int]:
        """Парсит значение ячейки в float, возвращает None если пусто."""
        if value is None:
            return None
        try:
            result = round(float(value))
            return result
        except (ValueError, TypeError):
            return None

    def _get_cell_color(self, cell) -> str:
        """
        Извлекает цвет заливки ячейки как строку для группировки.
        Возвращает hex-строку цвета или 'none' если без заливки.
        """
        fill = cell.fill
        if fill is None or fill.fgColor is None:
            return "none"

        fg = fill.fgColor

        if fg.type == "rgb" and fg.rgb is not None:
            rgb = str(fg.rgb)
            # openpyxl иногда возвращает ARGB (8 символов) — убираем альфа
            if len(rgb) == 8:
                return rgb[2:]
            return rgb
        elif fg.type == "theme":
            return f"theme_{fg.theme}_{fg.tint}"
        elif fg.type == "indexed":
            return f"indexed_{fg.indexed}"

        return "none"

    def _group_months_by_color(
        self,
        month_data: Dict[int, Dict]
    ) -> Dict[str, Dict[int, Dict]]:
        """
        Группирует месяцы по цвету ячеек → разные РК.

        :param month_data: {month: {"trp_total", "trp_20s", "color"}}
        :return: {color: {month: {"trp_total", "trp_20s"}}}
        """
        grouped = defaultdict(dict)

        for month, info in month_data.items():
            color = info["color"]
            grouped[color][month] = {
                "trp_total": info["trp_total"],
                "trp_20s": info["trp_20s"]
            }

        return dict(grouped)
    